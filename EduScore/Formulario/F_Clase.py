# Jose Antonio Garcia Peña (Diseño)
#
# 18/04/2023

import tkinter as tk
import util.generic as utl
from util.BotonRedondeado import RoundedButton
from tkinter import *
from tkinter.font import BOLD
from tkinter import ttk, messagebox
import os
import openpyxl
# from Formulario.F_Master import MasterPanel
# from Formulario.F_Registrar import Registrar_Panel

class Crear_Clases:
    
    #---------------
    # Funcion que permite la creacion de una nueva clase
    #---------------
    def crear(self):
        #---------------
        # Obtiene los valores necesarios para crear la clase
        #---------------
        Cantidad_Alumnos = self.Cant_Alumnos.get()
        Cantidad_Dias = self.Cant_Dias.get()
        Porcentaje_Examen = self.Por_Examen.get()
        Porcentaje_Tarea = self.Por_Tarea.get()
        Porcentaje_Asistencia = self.Por_Asistencia.get()
        Porcentaje_PF = self.Por_PF.get()
        Nom_archivo = self.Nombre_archivo.get()
        ruta_carpeta = "Docs"

        # Crear un nuevo libro de trabajo en Excel
        wb = openpyxl.Workbook()

        # Seleccionar la hoja activa
        ws = wb.active

        # Agregar los valores a las celdas especificadas
        ws.cell(row=2, column=1, value="Trabajos:")
        ws.cell(row=3, column=1, value="Examenes:")
        ws.cell(row=4, column=1, value="Proyecto Final:")
        ws.cell(row=5, column=1, value="Asistencias:")

        ws.cell(row=2, column=2, value=int(Porcentaje_Tarea))
        ws.cell(row=3, column=2, value=int(Porcentaje_Examen))
        ws.cell(row=4, column=2, value=int(Porcentaje_PF))
        ws.cell(row=5, column=2, value=int(Porcentaje_Asistencia))

        # Definir títulos de columnas
        column_titles = ["Apartados de evaluacion","Porcentajes de Evaluacion", "No.", "Alumnos", "Correos", "Calificación Final", "Trabajos", "Act 1", "Act 2", "Act 3", "Examen Total", "Examen medio", "Examen final", "Proyecto Final", "Asistencias"]
        for i in range(1, int(Cantidad_Dias) + 1):
            column_titles.append("Dia {}".format(i))

        # Escribir los títulos de columnas
        for col_num, title in enumerate(column_titles, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = title

        # Escribir la información de los alumnos
        for row_num in range(2, int(Cantidad_Alumnos) + 2):
            # Escribir número de fila en la primera columna
            cell = ws.cell(row=row_num, column=3)
            cell.value = row_num - 1

        # Guardar el libro de trabajo en Excel
        ruta_archivo = os.path.join(ruta_carpeta, Nom_archivo + ".xlsx")
        wb.save(ruta_archivo)
        self.ventana.destroy()

    #---------------
    # generar ventana.
    #---------------
    def __init__(self):
        self.ventana = tk.Tk()                             
        self.ventana.title('Creacion de clase')
        self.ventana.geometry('900x600')
        self.ventana.config(bg='#fcfcfc')
        self.ventana.resizable(width=0, height=0)  
        utl.centrar_ventana(self.ventana, 600, 400)
        
        #---------------
        # Frame principal de la ventana.
        #---------------
        frame_Clase = tk.Frame(self.ventana, bd = 0, relief= tk.SOLID, bg= '#FFFFFF')
        frame_Clase.pack(side = "top", expand= tk.YES, fill= tk.BOTH)
        #---------------
        # Frame contenedor del titulo.
        #---------------
        frame_title = tk.Frame(frame_Clase, bd= 0, height=60, relief= tk.SOLID, bg='#FFFFFF')
        frame_title.pack(side= "top", fill= tk.X)
        #---------------
        # Titulo.
        #---------------
        titulo = tk.Label(frame_title, bd= 0, text= "Crea una clase", font= ('kollektief', 48, 'bold'), fg= '#FFFFFF', bg= '#313745')
        titulo.pack(expand= tk.YES, fill= tk.BOTH)
        #---------------
        # Frame contenedor este contendra el resto de los widjes.
        #---------------
        frame_Contenedor = tk.Frame(frame_Clase, bd= 0, relief= tk.SOLID, bg= '#313745')
        frame_Contenedor.pack(side= 'bottom', expand= tk.YES, fill= tk.BOTH)
        #---------------
        # Etiqueta alumnos.
        #---------------
        label_alumnos = tk.Label(frame_Contenedor, text= "Cantidad de alumnos", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_alumnos.pack(fill= tk.X, padx= 70)
        label_alumnos.place(x= 70, y= 10)
        #---------------
        # Entry alumnos.
        #---------------
        self.Cant_Alumnos = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Cant_Alumnos.pack(fill= tk.X, padx= 70, pady= 10)
        self.Cant_Alumnos.place(x= 70, y= 40)
        #---------------
        # Etiqueta examen.
        #---------------
        label_Examen = tk.Label(frame_Contenedor, text= "Porcentaje de examen", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_Examen.pack(fill= tk.X, padx= 70)
        label_Examen.place(x= 70, y= 70)
        #---------------
        # Entry examen.
        #---------------
        self.Por_Examen = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Por_Examen.pack(fill= tk.X, padx= 70, pady= 10)
        self.Por_Examen.place(x= 70, y= 100)
        #---------------
        # Etiqueta tarea.
        #---------------
        label_Tarea = tk.Label(frame_Contenedor, text= "Porcentaje de tarea", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_Tarea.pack(fill= tk.X, padx= 70)
        label_Tarea.place(x= 70, y= 130)
        #---------------
        # Entry tarea.
        #---------------
        self.Por_Tarea = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Por_Tarea.pack(fill= tk.X, padx= 70, pady= 10)
        self.Por_Tarea.place(x= 70, y= 160)
        #---------------
        # Etiqueta Asistencia.
        #---------------
        label_Asistencia = tk.Label(frame_Contenedor, text= "Porcentaje por asistencia", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_Asistencia.pack(fill= tk.X, padx= 70)
        label_Asistencia.place(x= 350, y= 10)
        #---------------
        # Entry Asistencia.
        #---------------
        self.Por_Asistencia = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Por_Asistencia.pack(fill= tk.X, padx= 70, pady= 10)
        self.Por_Asistencia.place(x= 350, y= 40)
        #---------------
        # Etiqueta Asistencia, aproximacion de aistencia.
        #---------------
        label_Aprox_Asistencia = tk.Label(frame_Contenedor, text= "Dias aproximados del ciclo", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_Aprox_Asistencia.pack(fill= tk.X, padx= 70)
        label_Aprox_Asistencia.place(x= 350, y= 70)
        #---------------
        # Entry Asistencia, aproximacion de aistencia.
        #---------------
        self.Cant_Dias = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Cant_Dias.pack(fill= tk.X, padx= 70, pady= 10)
        self.Cant_Dias.place(x= 350, y= 100)
        #---------------
        # Etiqueta proyecto final.
        #---------------
        label_Aprox_Asistencia = tk.Label(frame_Contenedor, text= "Porcentaje por proyecto final", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_Aprox_Asistencia.pack(fill= tk.X, padx= 70)
        label_Aprox_Asistencia.place(x= 350, y= 130)
        #---------------
        # Entry proyecto final.
        #---------------
        self.Por_PF = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Por_PF.pack(fill= tk.X, padx= 70, pady= 10)
        self.Por_PF.place(x= 350, y= 160)
        #---------------
        # Etiqueta nombre del archivo.
        #---------------
        label_Aprox_Asistencia = tk.Label(frame_Contenedor, text= "Nombre de la clase", font=('Gidole Regular', 12), fg= '#FFFFFF', bg= '#313745', anchor= 'w')
        label_Aprox_Asistencia.pack(fill= tk.X, padx= 70)
        label_Aprox_Asistencia.place(x= 70, y= 190)
        #---------------
        # Entry nombre del archivo.
        #---------------
        self.Nombre_archivo = ttk.Entry(frame_Contenedor, font= ('Gidole Regular', 12))
        self.Nombre_archivo.pack(fill= tk.X, padx= 70, pady= 10)
        self.Nombre_archivo.place(x= 70, y= 220)
        #---------------
        # ejecuta la bentana.
        #---------------
        button_inicio = RoundedButton(frame_Contenedor, 140, 40, 5, 0, '#474E60', '#313745', text='Crear clase', command= self.crear)
        button_inicio.pack (side= 'right', padx= 70, pady= 0)
        button_inicio.place(x= 350, y= 220)
        #---------------
        # ejecuta la bentana.
        #---------------
        self.ventana.mainloop()


